from modulefinder import test

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from .models import Test, Question, Answer, Result


def web_login(request):
    if request.user.is_authenticated:
        return redirect('/web/dashboard/')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user and user.role in ('teacher', 'admin'):
            auth_login(request, user)
            return redirect('/web/dashboard/')
        else:
            error = 'Неверный логин/пароль или недостаточно прав'

    return render(request, 'api/login.html', {'error': error})
@login_required(login_url='/web/login/')
def result_review(request, result_id):
    from .models import UserAnswer
    result = get_object_or_404(
        Result.objects.select_related('test', 'user'),
        id=result_id, test__created_by=request.user
    )

    if request.method == 'POST':
        for ua in result.user_answers.filter(question__question_type='text'):
            grade_value = request.POST.get(f'grade_{ua.id}')
            if grade_value is not None:
                is_correct = grade_value == 'correct'
                ua.is_correct = is_correct
                ua.save()

        total = 0
        for ua in result.user_answers.all():
            if ua.is_correct:
                total += ua.question.points
        result.score = total
        result.save()
        return redirect(f'/web/results/review/{result.id}/?saved=1')

    user_answers = result.user_answers.select_related('question').prefetch_related(
        'selected_answers', 'question__answers'
    )

    return render(request, 'api/result_review.html', {
        'result': result,
        'user_answers': user_answers,
    })

def web_logout(request):
    auth_logout(request)
    return redirect('/web/login/')


@login_required(login_url='/web/login/')
def dashboard(request):
    tests = Test.objects.filter(created_by=request.user)
    results = Result.objects.filter(test__created_by=request.user, is_completed=True)
    return render(request, 'api/dashboard.html', {
        'tests_count': tests.count(),
        'published_count': tests.filter(is_published=True).count(),
        'results_count': results.count(),
        'recent_tests': tests.order_by('-created_at')[:5],
    })


@login_required(login_url='/web/login/')
def test_list(request):
    search_query = request.GET.get('q', '').strip()
    tests = Test.objects.filter(created_by=request.user).order_by('-created_at')

    if search_query:
        from django.db.models import Q
        tests = tests.filter(
            Q(title__icontains=search_query) | Q(subject__icontains=search_query)
        )

    return render(request, 'api/test_list.html', {
        'tests': tests,
        'search_query': search_query,
    })

@login_required(login_url='/web/login/')
def test_create(request):
    if request.method == 'POST':
        test = Test.objects.create(
    title=request.POST.get('title'),
    subject=request.POST.get('subject', ''),
    description=request.POST.get('description', ''),
    time_limit=int(request.POST.get('time_limit', 0)),
    attempts_allowed=int(request.POST.get('attempts_allowed', 1)),
    questions_per_attempt=int(request.POST.get('questions_per_attempt', 0)),
    shuffle_questions='shuffle_questions' in request.POST,
    shuffle_answers='shuffle_answers' in request.POST,
    created_by=request.user,
)
        return redirect(f'/web/tests/{test.id}/edit/')
    return render(request, 'api/test_create.html')


@login_required(login_url='/web/login/')
def test_edit(request, test_id):
    test = get_object_or_404(Test, id=test_id, created_by=request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'edit_test':
            test.title = request.POST.get('title')
            test.subject = request.POST.get('subject', '')
            test.description = request.POST.get('description', '')
            test.time_limit = int(request.POST.get('time_limit', 0))
            test.attempts_allowed = int(request.POST.get('attempts_allowed', 1))
            test.questions_per_attempt = int(request.POST.get('questions_per_attempt', 0))
            test.shuffle_questions = 'shuffle_questions' in request.POST
            test.shuffle_answers = 'shuffle_answers' in request.POST
            test.save()

        elif action == 'add_question':
            question_type = request.POST.get('question_type', 'single')
            question_text = request.POST.get('question_text', '').strip()

            if not question_text:
                return redirect(f'/web/tests/{test.id}/edit/')

            answers = []
            for i in range(1, 5):
                text = request.POST.get(f'answer_text_{i}', '').strip()
                if text:
                    is_correct = f'answer_correct_{i}' in request.POST
                    answers.append((text, is_correct))

            if question_type != 'text':
                if not answers:
                    return redirect(f'/web/tests/{test.id}/edit/?error=no_answers')

                correct_count = sum(1 for _, is_correct in answers if is_correct)

                if correct_count == 0:
                    return redirect(f'/web/tests/{test.id}/edit/?error=no_correct')

                if question_type == 'single' and correct_count > 1:
                    return redirect(f'/web/tests/{test.id}/edit/?error=too_many_correct')

            question = Question.objects.create(
                test=test,
                text=question_text,
                question_type=question_type,
                points=int(request.POST.get('points', 1)),
                order=test.questions.count(),
            )

            if question_type != 'text':
                for text, is_correct in answers:
                    Answer.objects.create(
                        question=question,
                        text=text,
                        is_correct=is_correct,
                    )

        elif action == 'delete_question':
            question_id = request.POST.get('question_id')
            Question.objects.filter(id=question_id, test=test).delete()

        return redirect(f'/web/tests/{test.id}/edit/')

    return render(request, 'api/test_edit.html', {'test': test})


@login_required(login_url='/web/login/')
def test_toggle(request, test_id):
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if request.method == 'POST':
        test.is_published = not test.is_published
        test.save()
    return redirect('/web/tests/')


@login_required(login_url='/web/login/')
def test_delete(request, test_id):
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if request.method == 'POST':
        test.delete()
    return redirect('/web/tests/')


@login_required(login_url='/web/login/')
def results(request):
    selected_test = request.GET.get('test_id', '')
    selected_username = request.GET.get('username', '')

    qs = Result.objects.filter(
        test__created_by=request.user, is_completed=True
    ).select_related('test', 'user').order_by('-finished_at')

    if selected_test:
        qs = qs.filter(test_id=selected_test)
    if selected_username:
        qs = qs.filter(user__username__icontains=selected_username)

    avg_percent = 0
    grade5_count = 0
    grade2_count = 0
    if qs.exists():
        percents = []
        for r in qs:
            if r.max_score > 0:
                percents.append(r.score / r.max_score * 100)
            grade = r.grade()
            if grade == 5:
                grade5_count += 1
            elif grade == 2:
                grade2_count += 1
        avg_percent = round(sum(percents) / len(percents), 1) if percents else 0

    tests = Test.objects.filter(created_by=request.user)
    return render(request, 'api/results.html', {
        'results': qs,
        'tests': tests,
        'selected_test': selected_test,
        'selected_username': selected_username,
        'avg_percent': avg_percent,
        'grade5_count': grade5_count,
        'grade2_count': grade2_count,
    })
def web_register(request):
    if request.user.is_authenticated:
        return redirect('/web/dashboard/')

    error = None
    form_data = {'username': '', 'email': '', 'role': 'student'}

    if request.method == 'POST':
        from .models import CustomUser

        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        role = request.POST.get('role', 'student')

        form_data = {'username': username, 'email': email, 'role': role}

        if not username or not email or not password:
            error = 'Заполните все поля'
        elif password != password_confirm:
            error = 'Пароли не совпадают'
        elif len(password) < 6:
            error = 'Пароль должен быть не короче 6 символов'
        elif CustomUser.objects.filter(username=username).exists():
            error = 'Пользователь с таким логином уже существует'
        elif CustomUser.objects.filter(email=email).exists():
            error = 'Пользователь с таким email уже существует'
        elif role not in ('student', 'teacher'):
            error = 'Недопустимая роль'
        else:
            user = CustomUser.objects.create_user(
                username=username, email=email, password=password, role=role
            )
            if role == 'teacher':
                auth_login(request, user)
                return redirect('/web/dashboard/')
            else:
                return redirect('/web/login/?registered=1')

    return render(request, 'api/register.html', {'error': error, 'form_data': form_data})

@login_required(login_url='/web/login/')
def test_statistics(request, test_id):
    from .models import UserAnswer
    test = get_object_or_404(Test, id=test_id, created_by=request.user)

    questions_stats = []
    for question in test.questions.all().order_by('order'):
        user_answers = UserAnswer.objects.filter(
            question=question,
            result__is_completed=True
        )
        total_attempts = user_answers.count()
        correct_count = user_answers.filter(is_correct=True).count()
        wrong_count = user_answers.filter(is_correct=False).count()
        pending_count = user_answers.filter(is_correct__isnull=True).count()

        if total_attempts > 0:
            success_rate = round(correct_count / total_attempts * 100, 1)
        else:
            success_rate = None

        # Уровень сложности
        if success_rate is None:
            difficulty = 'no_data'
            difficulty_label = 'Нет данных'
        elif success_rate >= 80:
            difficulty = 'easy'
            difficulty_label = 'Лёгкий'
        elif success_rate >= 50:
            difficulty = 'medium'
            difficulty_label = 'Средний'
        else:
            difficulty = 'hard'
            difficulty_label = 'Сложный'

        questions_stats.append({
            'question': question,
            'total': total_attempts,
            'correct': correct_count,
            'wrong': wrong_count,
            'pending': pending_count,
            'success_rate': success_rate,
            'difficulty': difficulty,
            'difficulty_label': difficulty_label,
        })

    # Общая статистика по тесту
    results_qs = test.results.filter(is_completed=True)
    total_attempts = results_qs.count()
    avg_score_percent = 0
    if total_attempts > 0 and test.questions.exists():
        total_max = sum(r.max_score for r in results_qs if r.max_score > 0)
        total_score = sum(r.score for r in results_qs if r.max_score > 0)
        if total_max > 0:
            avg_score_percent = round(total_score / total_max * 100, 1)

    return render(request, 'api/test_statistics.html', {
        'test': test,
        'questions_stats': questions_stats,
        'total_attempts': total_attempts,
        'avg_score_percent': avg_score_percent,
    })
@login_required(login_url='/web/login/')
def results_export(request):
    """Экспорт результатов в Excel с учётом фильтров."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    import datetime

    selected_test = request.GET.get('test_id', '')
    selected_username = request.GET.get('username', '')

    qs = Result.objects.filter(
        test__created_by=request.user, is_completed=True
    ).select_related('test', 'user').order_by('-finished_at')

    if selected_test:
        qs = qs.filter(test_id=selected_test)
    if selected_username:
        qs = qs.filter(user__username__icontains=selected_username)

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # Шапка
    headers = ['№', 'Студент', 'Тест', 'Балл', 'Макс. балл',
               'Процент', 'Оценка', 'Дата', 'Переключения окон']
    ws.append(headers)

    # Стиль шапки
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Данные
    for i, r in enumerate(qs, start=1):
        percent = round(r.score / r.max_score * 100, 1) if r.max_score > 0 else 0
        finished = r.finished_at.strftime('%d.%m.%Y %H:%M') if r.finished_at else ''
        ws.append([
            i, r.user.username, r.test.title,
            r.score, r.max_score, f"{percent}%", r.grade(),
            finished, r.tab_switches,
        ])

    # Ширина колонок
    widths = [5, 20, 30, 8, 12, 10, 8, 18, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Центрируем числовые колонки
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [1, 4, 5, 6, 7, 9]:
            row[col_idx - 1].alignment = center

    # Подсветка большого числа переключений
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[8]  # tab_switches
        if isinstance(cell.value, int) and cell.value > 0:
            cell.fill = red_fill
            cell.font = Font(bold=True, color='991B1B')

    # Отдаём файл
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"results_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response